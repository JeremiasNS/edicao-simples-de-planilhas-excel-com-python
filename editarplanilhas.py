#!/usr/bin/env python3.8

from openpyxl import load_workbook
#caminho_linux = '/media/jeremias/STORAGE/Documentos/Editando Tabelas com Python/Envio de Mensagens via Robo/cancelados.xlsx'

caminho_windows_ori = 'C:/Users/Hyago12/Documents/edicao-simples-de-planilhas-excel-com-python-master/cancelados.xlsx'
caminho_windows_dest = 'C:/Users/Hyago12/Documents/edicao-simples-de-planilhas-excel-com-python-master/destino.xlsx'

arquivo_excel_ori = load_workbook(caminho_windows_ori)
arquivo_excel_dest = load_workbook(caminho_windows_dest)

planilha1 = arquivo_excel_ori.active
planilha2 = arquivo_excel_dest.active

#Pede um número
numero = int(input("Digite um Número: "))

#Procura as posições equivalentes ao número informado
max_linha = planilha1.max_row
max_coluna = planilha1.max_column
for i in range(1, max_linha + 1):
    for j in range(1, max_coluna + 1):
        if(planilha1.cell(row=i, column=j).value == numero):
            linha = i 
            coluna = j
        
#print(planilha1.cell(row=i, column=j).value, end=" - ")

#a1 = planilha1['A1']#Nome
#b1 = planilha1['B1']#Telefone
## Imprime o valor da célula A1
#print("Nome: " + a1.value)
#print("Telefone: " + str(b1.value))
##print("A: " + str(b1.value)) 63992332727

#print("A linha: {}, e a coluna: {}, é onde se encontra o número: {}".format(linha,coluna,numero))
#Escrevo os dados na planilha de destino
#

max_row_plan2 = planilha2.max_row

dados = [];

#number = planilha1.cell(column=coluna, row=linha)
#nome = planilha1.cell(column=coluna-1, row=linha)

for j in range(0, max_coluna):
    print("j: "+str(j))
    dados[j] = planilha1.cell(j+1, linha)

for numb in dados:
    planilha2.cell( max_row_plan2, coluna, num)

#Retiro da planilha de origem
#planilha1.cell(linha, coluna, value="")
#planilha1.cell(linha, coluna-1, value="")
#dados[j] = planilha1.cell(column=max_coluna, row=linha)
#column=

#arquivo_excel_ori.save('cancelados.xlsx')
arquivo_excel_dest.save('destino.xlsx')
